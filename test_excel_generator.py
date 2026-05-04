"""
Phase 5 test for excel_generator.

NO API CALLS. Builds 3 simple SplitRows + 6 Meta-split SplitRows (the
TTTW5EZGT2 case from Phase 4), generates the workbook, then round-trips the
bytes through openpyxl and asserts every property finance relies on.

Coverage:
  a. Two sheets exist with the expected names.
  b. Header row contains the 20 documented column titles in order.
  c. Data rows reflect the input SplitRows in cell value, GL account,
     description, brand dimensions, etc.
  d. Freeze panes are set to "B2".
  e. Lookups sheet exposes every DROPDOWN_SEED category with its values.
  f. Data validations exist on columns C, F, G, H, I, J pointing at the
     right Lookups columns.
  g. Number formats are correct: SGD = "#,##0.00"; VND merchant amounts =
     "#,##0"; bank date = "DD-MMM-YYYY".
  h. The workbook round-trips through openpyxl without error (file is
     structurally valid and re-readable).
  i. Sheet name is sanitised + truncated.
  j. Auto-filter is set on the header range covering all data rows.
"""

from __future__ import annotations

import sys
from datetime import date
from io import BytesIO

import openpyxl

from config import DROPDOWN_SEED
from excel_generator import (
    COLUMNS,
    DROPDOWN_COLUMNS,
    NUMBER_FORMAT_DATE,
    NUMBER_FORMAT_SGD,
    NUMBER_FORMAT_VND,
    generate_output_excel,
)
from meta_splitter import split_meta_transaction
from models import (
    BankRow,
    MetaCampaign,
    MetaInvoiceOCR,
    SplitRow,
    TransactionType,
)


PASSED: list[str] = []
FAILED: list[str] = []


def check(label: str, ok: bool, detail: str = "") -> None:
    if ok:
        PASSED.append(label)
        print(f"  PASS  {label}")
    else:
        FAILED.append(f"{label} :: {detail}")
        print(f"  FAIL  {label}  ({detail})")


def section(t: str) -> None:
    print("\n" + "=" * 78)
    print(t)
    print("=" * 78)


# ---------------------------------------------------------------------------
# Fixtures: 3 simple + 6 Meta rows
# ---------------------------------------------------------------------------


def build_simple_rows() -> list[SplitRow]:
    """Three diverse simple rows: SGD Grab, VND Singapore Airlines, refund."""
    return [
        SplitRow(
            description="OCBC: PCard - Sharry - Grab - Booking#A-9XRJTDGGX9SFAV - "
                        "TPT in Singapore (SGD42.09)",
            gl_account="6312204",
            gl_account_name="Travelling - Sales Staff",
            line_number=1,
            brand="0_DIM2",
            country="SG",
            division="MED-I",
            team="0_DIM5",
            tax_code="",
            sgd_amount=42.09,
            bank_date=date(2026, 3, 1),
            bank_narrative="Grab* A-9XRJTDGGX9SFAV Singapore",
            card_last4="9804",
            cardholder="Sharry",
            merchant_currency="SGD",
            merchant_amount=42.09,
            bank_sgd=42.09,
            match_status="Matched",
            row_type="simple",
        ),
        SplitRow(
            description="OCBC: PCard - Sharry - Singapore Airlines - "
                        "Booking#DDOWAT - Air-ticket 05/03/26, SGN/SIN (VND3,681,000)",
            gl_account="6312204",
            gl_account_name="Travelling - Sales Staff",
            line_number=2,
            brand="0_DIM2",
            country="VN",
            division="0_DIM4",
            team="0_DIM5",
            tax_code="ZP",
            sgd_amount=186.10,
            bank_date=date(2026, 2, 4),
            bank_narrative="SQ INTERNET PURCHASES SINGAPORE",
            card_last4="9804",
            cardholder="Sharry",
            merchant_currency="VND",
            merchant_amount=3_681_000,
            bank_sgd=186.10,
            match_status="Matched",
            row_type="simple",
        ),
        SplitRow(
            description="OCBC: PCard - KC - TAOBAO Singapore (SGD-77.27)",
            gl_account=None,
            gl_account_name=None,
            line_number=3,
            brand="0_DIM2",
            country="",
            division="0_DIM4",
            team="0_DIM5",
            tax_code="",
            sgd_amount=-77.27,
            bank_date=date(2026, 3, 2),
            bank_narrative="TAOBAO Singapore",
            card_last4="0711",
            cardholder="KC",
            merchant_currency="SGD",
            merchant_amount=-77.27,
            bank_sgd=-77.27,
            match_status="Unmatched",
            row_type="simple",
        ),
    ]


def build_meta_rows() -> list[SplitRow]:
    """Re-use the splitter to produce the 6 TTTW5EZGT2 rows from Phase 4."""
    bank = BankRow(
        row_index=10,
        sheet_name="Bank statement - Feb26",
        transaction_id="TX-TTTW5EZGT2",
        transaction_date=date(2026, 2, 14),
        narrative="FACEBK *TTTW5EZGT2 DUBLIN",
        last_4_digits="9804",
        cardholder="Sharry",
        merchant_amount=192508.0,
        merchant_currency="VND",
        sgd_amount=9.65,
        transaction_type=TransactionType.META,
        fb_code="TTTW5EZGT2",
        is_refund=False,
    )
    ocr = MetaInvoiceOCR(
        reference_number="TTTW5EZGT2",
        invoice_date=date(2026, 2, 14),
        payment_method_last4="9804",
        total_paid_vnd=192508,
        subtotal_vnd=175007,
        vat_vnd=17501,
        vat_rate_percent=10.0,
        campaigns=[
            MetaCampaign(ad_set_name="Calecim Brand Ads", spend_vnd=3568),
            MetaCampaign(ad_set_name="Heliocare Brands Campaign", spend_vnd=70042),
            MetaCampaign(ad_set_name="64MG Profhilo Webinar 28 Feb", spend_vnd=101397),
        ],
        source_file="TTTW5EZGT2.pdf",
    )
    return split_meta_transaction(bank, ocr, start_line=4)


# ---------------------------------------------------------------------------
# Run
# ---------------------------------------------------------------------------


def main() -> int:
    section("Build sample SplitRows")
    rows = build_simple_rows() + build_meta_rows()
    print(f"  Built {len(rows)} rows ({len(build_simple_rows())} simple + {len(build_meta_rows())} meta)")

    section("Generate Excel bytes")
    blob = generate_output_excel(rows, "Test")
    print(f"  Generated {len(blob):,} bytes")

    # ----- Round-trip via openpyxl ------------------------------------
    try:
        wb = openpyxl.load_workbook(BytesIO(blob))
    except Exception as exc:
        check("Workbook re-opens with openpyxl", False, str(exc))
        return 1
    check("Workbook re-opens with openpyxl", True)

    # (a) Sheets ------------------------------------------------------
    section("(a) Sheets")
    check("'PCard-Test' sheet present",
          "PCard-Test" in wb.sheetnames,
          f"got {wb.sheetnames}")
    check("'Lookups' sheet present",
          "Lookups" in wb.sheetnames,
          f"got {wb.sheetnames}")
    ws = wb["PCard-Test"]
    lk = wb["Lookups"]

    # (b) Header row -------------------------------------------------
    section("(b) Header row")
    actual_headers = [ws.cell(1, c).value for c, _, _ in COLUMNS]
    # openpyxl serialises an empty-string cell back as None on re-read, so
    # the spacer column (12, header "") shows up as None. Normalise both
    # sides for the comparison.
    expected_headers = [(h or None) for _, h, _ in COLUMNS]
    check("All 20 headers present in correct order",
          actual_headers == expected_headers,
          f"got {actual_headers}")
    # Spot-check a few critical ones
    check("Column B header == 'Description'",
          ws.cell(1, 2).value == "Description")
    check("Column K header == 'Unit Price'",
          ws.cell(1, 11).value == "Unit Price")
    check("Column T header == 'Match Status'",
          ws.cell(1, 20).value == "Match Status")

    # (c) Data rows ---------------------------------------------------
    section("(c) Data rows match SplitRows")
    check(f"{len(rows)} data rows present (max_row={ws.max_row})",
          ws.max_row == len(rows) + 1,
          f"got max_row={ws.max_row}")

    # Spot-check the simple Grab row (excel row 2)
    r = 2
    check("Row 2: '#' == 1", ws.cell(r, 1).value == 1)
    check("Row 2: description matches",
          ws.cell(r, 2).value == rows[0].description,
          f"got {ws.cell(r, 2).value!r}")
    check("Row 2: GL account == '6312204'",
          ws.cell(r, 3).value == "6312204")
    check("Row 2: Brand == '0_DIM2'",
          ws.cell(r, 6).value == "0_DIM2")
    check("Row 2: Card == '9804'",
          ws.cell(r, 15).value == "9804")
    check("Row 2: Cardholder == 'Sharry'",
          ws.cell(r, 16).value == "Sharry")
    check("Row 2: SGD Amount (col S) == 42.09",
          ws.cell(r, 19).value == 42.09)
    check("Row 2: Match Status == 'Matched'",
          ws.cell(r, 20).value == "Matched")

    # Spot-check the Meta Calecim spend row (first Meta row, excel row 5)
    r = 5
    check("Row 5: description references TTTW5EZGT2",
          "TTTW5EZGT2" in str(ws.cell(r, 2).value),
          f"got {ws.cell(r, 2).value!r}")
    check("Row 5: GL account == '6210101' (Advertisement)",
          ws.cell(r, 3).value == "6210101")
    check("Row 5: Brand == 'CAL'",
          ws.cell(r, 6).value == "CAL")
    check("Row 5: Division == 'MED-I'",
          ws.cell(r, 8).value == "MED-I")
    check("Row 5: Match Status == 'Meta Split'",
          ws.cell(r, 20).value == "Meta Split")

    # Penny-perfect: SGD amounts on the 6 Meta rows sum to 9.65
    meta_rows_excel = list(range(5, 11))  # rows 5..10
    meta_unit_prices = [ws.cell(r, 11).value for r in meta_rows_excel]
    from decimal import Decimal
    sgd_sum = sum(Decimal(str(v)) for v in meta_unit_prices)
    check("Meta unit prices in Excel sum to exactly 9.65 (penny-perfect)",
          sgd_sum == Decimal("9.65"),
          f"got {sgd_sum}")

    # (d) Freeze panes ------------------------------------------------
    section("(d) Freeze panes")
    check("freeze_panes == 'B2'",
          ws.freeze_panes == "B2",
          f"got {ws.freeze_panes!r}")

    # (e) Lookups sheet -----------------------------------------------
    section("(e) Lookups sheet has dropdown seeds")
    keys = list(DROPDOWN_SEED.keys())
    actual_lk_headers = [lk.cell(1, c).value for c in range(1, len(keys) + 1)]
    check("Lookups headers match DROPDOWN_SEED keys in order",
          actual_lk_headers == keys,
          f"got {actual_lk_headers}")
    # Spot-check a couple of values
    gl_col = keys.index("GL_Account") + 1
    gl_values = [lk.cell(r, gl_col).value
                 for r in range(2, 2 + len(DROPDOWN_SEED["GL_Account"]))]
    check("Lookups: GL_Account column has the seeded values",
          gl_values == DROPDOWN_SEED["GL_Account"],
          f"got {gl_values}")
    brand_col = keys.index("Brand") + 1
    brand_values = [lk.cell(r, brand_col).value
                    for r in range(2, 2 + len(DROPDOWN_SEED["Brand"]))]
    check("Lookups: Brand column has the seeded values",
          brand_values == DROPDOWN_SEED["Brand"],
          f"got {brand_values}")

    # (f) Data validations on the right columns ----------------------
    section("(f) Data validations on columns C/F/G/H/I/J")
    validations = list(ws.data_validations.dataValidation)
    expected_columns = {c for c, _ in DROPDOWN_COLUMNS}
    found_columns: set[int] = set()
    found_keys: set[str] = set()

    from openpyxl.utils import column_index_from_string
    for dv in validations:
        # dv.sqref is a string or MultiCellRange like "C2:C10"
        for sqref in str(dv.sqref).split():
            col_letters = "".join(ch for ch in sqref.split(":")[0] if ch.isalpha())
            if col_letters:
                found_columns.add(column_index_from_string(col_letters))
        # Sanity: each formula references the Lookups sheet
        f1 = (dv.formula1 or "").replace("=", "")
        if f1.startswith("Lookups!"):
            for k in DROPDOWN_SEED:
                if f"!${chr(ord('A') + list(DROPDOWN_SEED).index(k))}$2:" in f1:
                    found_keys.add(k)

    check(f"Validations cover all {len(expected_columns)} expected columns",
          expected_columns.issubset(found_columns),
          f"missing={expected_columns - found_columns}, got {sorted(found_columns)}")
    check("All validation formulas reference the Lookups sheet",
          all((dv.formula1 or "").startswith("=Lookups!") for dv in validations),
          f"formulas={[dv.formula1 for dv in validations]}")
    check("All 6 dropdown lookup keys are wired up",
          found_keys == set(k for _, k in DROPDOWN_COLUMNS),
          f"got {sorted(found_keys)}")

    # (g) Number formats ---------------------------------------------
    section("(g) Number formats")
    # Row 2 (Grab, SGD): col 18 (merchant amount) should be SGD-style 2dp.
    check("Row 2 (SGD merchant) col R has SGD format",
          ws.cell(2, 18).number_format == NUMBER_FORMAT_SGD,
          f"got {ws.cell(2, 18).number_format!r}")
    # Row 3 (Singapore Airlines, VND merchant): col 18 should be VND integer.
    check("Row 3 (VND merchant) col R has VND format",
          ws.cell(3, 18).number_format == NUMBER_FORMAT_VND,
          f"got {ws.cell(3, 18).number_format!r}")
    # Unit Price (col K) and SGD Amount (col S) always SGD-style on every data row
    for r in range(2, len(rows) + 2):
        if ws.cell(r, 11).number_format != NUMBER_FORMAT_SGD:
            check(f"Row {r}: col K (Unit Price) has SGD format",
                  False, f"got {ws.cell(r, 11).number_format!r}")
            break
    else:
        check("All data rows: col K (Unit Price) has SGD format", True)
    for r in range(2, len(rows) + 2):
        if ws.cell(r, 19).number_format != NUMBER_FORMAT_SGD:
            check(f"Row {r}: col S (SGD Amount) has SGD format",
                  False, f"got {ws.cell(r, 19).number_format!r}")
            break
    else:
        check("All data rows: col S (SGD Amount) has SGD format", True)
    check("Bank Date col M uses date format",
          ws.cell(2, 13).number_format == NUMBER_FORMAT_DATE,
          f"got {ws.cell(2, 13).number_format!r}")

    # (j) Auto-filter -------------------------------------------------
    section("(j) Auto-filter")
    check(f"auto_filter.ref covers A1:T{len(rows) + 1}",
          ws.auto_filter.ref == f"A1:T{len(rows) + 1}",
          f"got {ws.auto_filter.ref!r}")

    # (i) Sheet-name sanitisation (separate test, doesn't need round-trip) -
    section("(i) Sheet-name sanitisation")
    from excel_generator import _safe_sheet_name
    check("Strips invalid chars [/\\?*]",
          "/" not in _safe_sheet_name("PCard/Test*?"),
          f"got {_safe_sheet_name('PCard/Test*?')!r}")
    check("Truncates to 31 chars",
          len(_safe_sheet_name("a" * 100)) == 31)
    check("Falls back when input is empty",
          _safe_sheet_name("") == "Sheet")

    section("Result")
    print(f"  PASSED: {len(PASSED)}")
    print(f"  FAILED: {len(FAILED)}")
    if FAILED:
        print("\nFailures:")
        for f in FAILED:
            print(f"  - {f}")
        return 1
    print("\nPhase 5 Excel generator test passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
