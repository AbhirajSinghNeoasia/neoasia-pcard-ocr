r"""
SAP-journal-entry Excel generator.

Produces an .xlsx workbook (returned as bytes — generated entirely in-memory
so it can be served straight from a Streamlit download button) containing two
sheets:

  Sheet 1: "PCard-{statement_period}" — the journal entry rows
           Columns A..T per CLAUDE.md Section 5.
  Sheet 2: "Lookups"                  — dropdown source values

Visual / behavioural spec (Section 5):
  - Header row: navy (#004D71) fill, white Calibri 11 bold.
  - Data rows: alternating white / very light blue (#E7EFF8).
  - Borders: thin, light gray (#EAEAEA).
  - Number formats: SGD = "#,##0.00"; VND merchant amounts = "#,##0".
  - Date format: "DD-MMM-YYYY".
  - Freeze panes at B2 (row 1 + column A locked).
  - Auto-filter on the header range.
  - Data validations (dropdowns) on columns C / F / G / H / I / J pointing at
    the corresponding column in the Lookups sheet.
  - Sheet names sanitised for Excel's invalid-character set ([]:*?/\) and
    truncated to Excel's 31-character limit.
"""

from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from config import COLORS, DROPDOWN_SEED, EXCEL_FONT
from models import SplitRow


# ---------------------------------------------------------------------------
# Column layout — single source of truth
# ---------------------------------------------------------------------------

# (col_index_1based, header, width)
COLUMNS: list[tuple[int, str, int]] = [
    (1,  "#",                  5),
    (2,  "Description",        60),
    (3,  "G/L Account",        12),
    (4,  "G/L Account Name",   25),
    (5,  "Line Number",        8),
    (6,  "Brand",              10),
    (7,  "Country",            8),
    (8,  "Division",           10),
    (9,  "Sales Team",         10),
    (10, "Tax Code",           8),
    (11, "Unit Price",         14),
    (12, "",                   2),    # spacer
    (13, "Bank Date",          12),
    (14, "Bank Narrative",     35),
    (15, "Card",               6),
    (16, "Cardholder",         12),
    (17, "Merchant Currency",  8),
    (18, "Merchant Amount",    14),
    (19, "SGD Amount",         14),
    (20, "Match Status",       12),
]

LAST_COL = max(c for c, _, _ in COLUMNS)
LAST_COL_LETTER = get_column_letter(LAST_COL)

# Columns that get a dropdown (see CLAUDE.md Section 5).
# (col_index, lookup_seed_key)
DROPDOWN_COLUMNS: list[tuple[int, str]] = [
    (3,  "GL_Account"),
    (6,  "Brand"),
    (7,  "Country"),
    (8,  "Division"),
    (9,  "Team"),
    (10, "Tax_Code"),
]

NUMBER_FORMAT_SGD = "#,##0.00"
NUMBER_FORMAT_VND = "#,##0"
NUMBER_FORMAT_DATE = "DD-MMM-YYYY"

INVALID_SHEET_CHARS = set('[]:*?/\\')
SHEET_NAME_MAX = 31


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def generate_output_excel(
    split_rows: Iterable[SplitRow],
    statement_period: str,
) -> bytes:
    """Build the SAP journal Excel and return the raw bytes."""
    rows = list(split_rows)

    wb = Workbook()
    main_ws = wb.active
    main_ws.title = _safe_sheet_name(f"PCard-{statement_period}")

    _write_header(main_ws)
    _write_data_rows(main_ws, rows)
    _apply_widths(main_ws)
    _freeze_and_filter(main_ws, last_data_row=len(rows) + 1)
    _apply_dropdowns(main_ws, last_data_row=len(rows) + 1)

    _write_lookups_sheet(wb)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Styles (built once and reused — PatternFill/Border are cheap to share)
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(
    start_color="FF" + COLORS["primary_dark"].lstrip("#").upper(),
    end_color="FF" + COLORS["primary_dark"].lstrip("#").upper(),
    fill_type="solid",
)
_HEADER_FONT = Font(name=EXCEL_FONT, size=11, bold=True, color="FFFFFFFF")
_HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)

_ALT_FILL = PatternFill(
    start_color="FF" + COLORS["very_light_blue"].lstrip("#").upper(),
    end_color="FF" + COLORS["very_light_blue"].lstrip("#").upper(),
    fill_type="solid",
)

_DATA_FONT = Font(name=EXCEL_FONT, size=10)

_BORDER_SIDE = Side(style="thin", color="FF" + COLORS["light_gray"].lstrip("#").upper())
_BORDER = Border(left=_BORDER_SIDE, right=_BORDER_SIDE, top=_BORDER_SIDE, bottom=_BORDER_SIDE)


# ---------------------------------------------------------------------------
# Header / widths / freeze / filter
# ---------------------------------------------------------------------------


def _write_header(ws) -> None:
    for col_idx, header, _w in COLUMNS:
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _BORDER
    ws.row_dimensions[1].height = 28


def _apply_widths(ws) -> None:
    for col_idx, _h, width in COLUMNS:
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _freeze_and_filter(ws, last_data_row: int) -> None:
    ws.freeze_panes = "B2"
    # Auto-filter on the header range (covering all data rows).
    end_row = max(1, last_data_row)
    ws.auto_filter.ref = f"A1:{LAST_COL_LETTER}{end_row}"


# ---------------------------------------------------------------------------
# Data rows
# ---------------------------------------------------------------------------


def _write_data_rows(ws, rows: list[SplitRow]) -> None:
    for i, sr in enumerate(rows):
        excel_row = i + 2  # row 1 is the header
        is_alt = (i % 2 == 1)

        cells: list[tuple[int, object]] = [
            (1,  i + 1),                                # # (sequential journal line)
            (2,  sr.description),                       # Description
            (3,  sr.gl_account or ""),                  # G/L Account
            (4,  sr.gl_account_name or ""),             # G/L Account Name
            (5,  sr.line_number),                       # Line Number
            (6,  sr.brand),                             # Brand
            (7,  sr.country),                           # Country
            (8,  sr.division),                          # Division
            (9,  sr.team),                              # Sales Team
            (10, sr.tax_code or ""),                    # Tax Code
            (11, sr.sgd_amount),                        # Unit Price (SGD)
            (12, ""),                                   # spacer
            (13, sr.bank_date),                         # Bank Date
            (14, sr.bank_narrative),                    # Bank Narrative
            (15, sr.card_last4),                        # Card
            (16, sr.cardholder),                        # Cardholder
            (17, sr.merchant_currency),                 # Merchant Currency
            (18, sr.merchant_amount),                   # Merchant Amount
            (19, sr.bank_sgd),                          # SGD Amount (bank charge)
            (20, sr.match_status),                      # Match Status
        ]

        for col_idx, value in cells:
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.value = value
            cell.font = _DATA_FONT
            cell.border = _BORDER
            if is_alt:
                cell.fill = _ALT_FILL

        # Per-cell number formats
        ws.cell(excel_row, 11).number_format = NUMBER_FORMAT_SGD       # Unit Price
        ws.cell(excel_row, 19).number_format = NUMBER_FORMAT_SGD       # SGD Amount
        if (sr.merchant_currency or "").upper() == "VND":
            ws.cell(excel_row, 18).number_format = NUMBER_FORMAT_VND   # Merchant Amount (VND)
        else:
            ws.cell(excel_row, 18).number_format = NUMBER_FORMAT_SGD
        if sr.bank_date is not None:
            ws.cell(excel_row, 13).number_format = NUMBER_FORMAT_DATE  # Bank Date


# ---------------------------------------------------------------------------
# Dropdowns + Lookups sheet
# ---------------------------------------------------------------------------


def _write_lookups_sheet(wb: Workbook) -> None:
    """Materialise DROPDOWN_SEED into a 'Lookups' sheet, one column per category."""
    ws = wb.create_sheet("Lookups")
    keys = list(DROPDOWN_SEED.keys())
    for col_idx, key in enumerate(keys, start=1):
        ws.cell(row=1, column=col_idx, value=key).font = Font(name=EXCEL_FONT, bold=True)
        for row_idx, value in enumerate(DROPDOWN_SEED[key], start=2):
            ws.cell(row=row_idx, column=col_idx, value=value)
        ws.column_dimensions[get_column_letter(col_idx)].width = 16


def _apply_dropdowns(ws, last_data_row: int) -> None:
    """Wire each dropdown column on the main sheet to the Lookups column."""
    if last_data_row < 2:
        return  # No data rows yet — skip (validations would target empty range)

    keys = list(DROPDOWN_SEED.keys())
    for col_idx, lookup_key in DROPDOWN_COLUMNS:
        if lookup_key not in keys:
            continue  # defensive
        lookup_col_letter = get_column_letter(keys.index(lookup_key) + 1)
        n_values = len(DROPDOWN_SEED[lookup_key])
        # Reference cells $A$2:$A$N on Lookups sheet.
        formula = f"=Lookups!${lookup_col_letter}$2:${lookup_col_letter}${n_values + 1}"
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Please pick a value from the dropdown list."
        dv.errorTitle = "Invalid value"
        ws.add_data_validation(dv)
        col_letter = get_column_letter(col_idx)
        dv.add(f"{col_letter}2:{col_letter}{last_data_row}")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _safe_sheet_name(name: str) -> str:
    cleaned = "".join(c if c not in INVALID_SHEET_CHARS else "-" for c in name)
    cleaned = cleaned.strip().strip("'") or "Sheet"
    return cleaned[:SHEET_NAME_MAX]
